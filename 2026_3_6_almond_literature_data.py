"""
almond_literature_data.py
=========================
Almond irrigation–yield production function:
  1. Hardcoded literature data (unit-corrected at source)
  2. Excel output (All_Observations, Study_Summaries, Parameter_Extraction)
  3. Step-by-step weighted parameter extraction → Bayesian priors

Unit corrections applied at source
-----------------------------------
  Naor 2017      : original kg/ha  → divided by 10 → kg/dunam
  Goldhamer 2005 : original kg/ha  → divided by 10 → kg/dunam
  Egea 2012      : original 6-year cumulative kg/dunam → divided by 6 → annual kg/dunam

Excluded from parameter fitting (kept in dataset, flagged)
-----------------------------------------------------------
  Miras 2023  : meta-analysis of ~15 Spanish studies — would double-count Spain
  Egea 2012   : young orchard (6-yr establishment) — tree age confounds water effect
"""

import numpy as np
import pandas as pd
from scipy.optimize import curve_fit
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — RAW DATA (all yields in kg/dunam/year, irrigation in m³/dunam)
# ─────────────────────────────────────────────────────────────────────────────

# Each row: [Irrigation_m3_per_dunam, Yield_kg_per_dunam, Rainfall_mm, Cultivar, Study, Notes]

# ── Miras 2023  |  Spain (meta-analysis)  |  EXCLUDED from fitting ──────────
# Meta-analysis of ~15 Spanish studies. Points already aggregate across experiments.
# Including alongside primary studies would double-count Spanish data.
miras_2023 = [
    [50,   50,  200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [100,  70,  200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [200,  100, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [300,  130, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [400,  160, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [500,  190, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [600,  215, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [700,  235, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [800,  250, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [900,  260, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [1000, 270, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
    [1050, 270, 200, 'Multiple', 'Miras 2023', 'Meta-analysis point'],
]

# ── Egea 2010  |  Murcia, Spain  |  Relevance = 2 ────────────────────────────
# PRD and RDI treatments; 3 seasons; cv Marta
# Units: kernel yield kg/dunam/year — no correction needed
egea_2010 = [
    [168, 101, 206, 'Marta', 'Egea 2010', 'PRD/RDI treatment 1'],
    [256, 118, 206, 'Marta', 'Egea 2010', 'PRD/RDI treatment 2'],
    [332, 121, 206, 'Marta', 'Egea 2010', 'PRD/RDI treatment 3'],
    [358, 128, 206, 'Marta', 'Egea 2010', 'PRD/RDI treatment 4'],
    [600, 144, 206, 'Marta', 'Egea 2010', 'Full irrigation control'],
]

# ── Egea 2012  |  Cartagena, Spain  |  EXCLUDED from fitting ─────────────────
# Original: cumulative kernel yield per tree over 6-year trial × 24 trees/dunam
# Correction: ÷6 to convert 6-year cumulative → annual average
# Excluded because orchard was in establishment phase — tree age confounds water effect
_egea2012_raw = [
    [165, 3044/6*10/10, 355, 'Marta', 'Egea 2012', 'SDI_ms - 37% of FI water'],
    [239, 3632/6*10/10, 355, 'Marta', 'Egea 2012', 'RDI - 55% of FI water'],
    [283, 3538/6*10/10, 355, 'Marta', 'Egea 2012', 'SDI_mm - 63% of FI water'],
    [446, 4318/6*10/10, 355, 'Marta', 'Egea 2012', 'FI - Full irrigation'],
]
# Explanation: original values were kg/tree × 24 trees/dunam (=×10/10, net ×1 after ÷10)
# then ÷6 for annual. Simplified: original stored values were already ×24 trees,
# so we just need ÷6 from what was in Excel.
egea_2012 = [
    [165,  round(507.333/6, 2), 355, 'Marta', 'Egea 2012', 'SDI_ms 37% FI | ÷6 cumul→annual'],
    [239,  round(605.333/6, 2), 355, 'Marta', 'Egea 2012', 'RDI 55% FI | ÷6 cumul→annual'],
    [283,  round(589.667/6, 2), 355, 'Marta', 'Egea 2012', 'SDI_mm 63% FI | ÷6 cumul→annual'],
    [446,  round(719.667/6, 2), 355, 'Marta', 'Egea 2012', 'FI full irrig | ÷6 cumul→annual'],
]

# ── Girona 2005  |  Catalonia, Spain  |  Relevance = 2 ──────────────────────
# RDI during kernel-filling; 4-year study; cv Ferragnès
# Units: kernel yield kg/dunam/year — no correction needed
girona_2005 = [
    [236, 141, 274, 'Ferragnès', 'Girona 2005', 'RDI treatment 1'],
    [399, 148, 274, 'Ferragnès', 'Girona 2005', 'RDI treatment 2'],
    [608, 176, 274, 'Ferragnès', 'Girona 2005', 'RDI treatment 3'],
    [799, 156, 274, 'Ferragnès', 'Girona 2005', 'RDI treatment 4'],
]

# ── Lopez 2018  |  Cordoba, Spain  |  Relevance = 2 ─────────────────────────
# RDI and SDI comparison; 3-year study; cv Guara; WP_ET = 0.23 kg/m³ reported
# Units: kernel yield kg/dunam/year — no correction needed
lopez_2018 = [
    [326, 150, 243, 'Guara', 'Lopez 2018', 'SDI treatment'],
    [504, 215, 243, 'Guara', 'Lopez 2018', 'RDI treatment 1'],
    [526, 204, 243, 'Guara', 'Lopez 2018', 'RDI treatment 2'],
    [754, 244, 243, 'Guara', 'Lopez 2018', 'Full irrigation control'],
]

# ── Moldero 2021  |  Cordoba, Spain  |  Relevance = 2 ───────────────────────
# Long-term DI; 4 treatments; 6-year study; cv Guara
# Units: kernel yield kg/dunam/year (converted from kg/ha ÷10 in original paper)
moldero_2021 = [
    [243, round(1430/10, 1), 238, 'Guara', 'Moldero 2021', 'RDI_S - 29% of FI water'],
    [524, round(2278/10, 1), 238, 'Guara', 'Moldero 2021', 'RDI_M - 62% of FI water'],
    [566, round(2339/10, 1), 238, 'Guara', 'Moldero 2021', 'SDI_M - 67% of FI water'],
    [840, round(2659/10, 1), 238, 'Guara', 'Moldero 2021', 'FI - Full irrigation'],
]

# ── Goldhamer 2005  |  California, USA  |  Relevance = 3 ────────────────────
# RDI timing study; 3 irrigation levels × 3 stress patterns; cv Nonpareil
# CORRECTION: original in kg/ha → ÷10 → kg/dunam
goldhamer_2005 = [
    [567, round(1754/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '580A Pre-harvest stress | ÷10 ha→dunam'],
    [580, round(1480/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '580B Post-harvest stress | ÷10 ha→dunam'],
    [610, round(1827/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '580C Sustained deficit | ÷10 ha→dunam'],
    [695, round(1849/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '720A Pre-harvest stress | ÷10 ha→dunam'],
    [702, round(1757/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '720B Post-harvest stress | ÷10 ha→dunam'],
    [762, round(1911/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '720C Sustained deficit | ÷10 ha→dunam'],
    [800, round(1839/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '860A Pre-harvest stress | ÷10 ha→dunam'],
    [888, round(1951/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '860B Post-harvest stress | ÷10 ha→dunam'],
    [892, round(2008/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', '860C Sustained deficit | ÷10 ha→dunam'],
    [976, round(2099/10, 1), 200, 'Nonpareil', 'Goldhamer 2005', 'Control 100% ETc | ÷10 ha→dunam'],
]

# ── Goldhamer 2016  |  California, USA  |  Relevance = 3 ────────────────────
# 5-year study; 10 irrigation levels (997–1341 m³/dunam); broadest water range
# Units: kernel yield kg/dunam/year — no correction needed
goldhamer_2016 = [
    [997,  330, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1076, 350, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1096, 360, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1146, 370, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1176, 380, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1216, 385, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1259, 390, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1320, 388, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1329, 387, 60, 'Nonpareil', 'Goldhamer 2016', ''],
    [1341, 385, 60, 'Nonpareil', 'Goldhamer 2016', ''],
]

# ── Naor 2017  |  Israel  |  Relevance = 5 ──────────────────────────────────
# Israeli cv Um-El-Fahem; GF677 rootstock; ~500 mm winter rain; 5 irrigation levels
# CORRECTION: original in kg/ha → ÷10 → kg/dunam
naor_2017 = [
    [394, round(1000/10, 1), 500, 'UmElFahem', 'Naor 2017', 'Low irrig. estimated from Fig 4a | ÷10 ha→dunam'],
    [413, round(1400/10, 1), 500, 'UmElFahem', 'Naor 2017', 'Low-medium | ÷10 ha→dunam'],
    [467, round(2000/10, 1), 500, 'UmElFahem', 'Naor 2017', 'Medium | ÷10 ha→dunam'],
    [605, round(2500/10, 1), 500, 'UmElFahem', 'Naor 2017', 'Medium-high | ÷10 ha→dunam'],
    [801, round(2800/10, 1), 500, 'UmElFahem', 'Naor 2017', 'High | ÷10 ha→dunam'],
]

# ── Sperling 2025  |  Israel (Lavi)  |  Relevance = 5 ──────────────────────
# Kamai lab; mechanistic soil-tree model; ~500 mm winter rain
# 3 irrigation treatments; yields estimated from Fig 4C
# Units: kernel yield kg/dunam/year — no correction needed
sperling_2025 = [
    [600,  180, 500, 'UmElFahem', 'Sperling 2025', 'Deficit irrig. — estimated from Fig 4C'],
    [900,  250, 500, 'UmElFahem', 'Sperling 2025', 'Moderate irrig. — estimated from Fig 4C'],
    [1250, 280, 500, 'UmElFahem', 'Sperling 2025', 'High irrig. — estimated from Fig 4C'],
]

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — ASSEMBLE DATAFRAME
# ─────────────────────────────────────────────────────────────────────────────

COLS = ['Irrigation_m3_per_dunam', 'Yield_kg_per_dunam',
        'Seasonal_Rainfall_mm', 'Cultivar', 'Study', 'Notes']

EXCLUDE_FROM_FIT = {'Miras 2023', 'Egea 2012'}

# Metadata: region, relevance score, exclusion flag, reason
STUDY_META = {
    'Naor 2017':      {'region': 'Israel',     'rel': 5, 'exclude': False,
                       'note': 'Israeli cv Um-El-Fahem, GF677 rootstock; arid conditions similar to Arad Valley'},
    'Sperling 2025':  {'region': 'Israel',     'rel': 5, 'exclude': False,
                       'note': 'Kamai lab; mechanistic soil-tree model; closest analogue to study farm'},
    'Goldhamer 2016': {'region': 'California', 'rel': 3, 'exclude': False,
                       'note': '5-yr; 10 irrig. levels; broadest water range; low rainfall (60 mm)'},
    'Goldhamer 2005': {'region': 'California', 'rel': 3, 'exclude': False,
                       'note': 'RDI timing; 10 treatments; Nonpareil cv; ÷10 unit correction applied'},
    'Moldero 2021':   {'region': 'Spain',      'rel': 2, 'exclude': False,
                       'note': '6-yr; 4 treatments; no exhaustion/adaptation detected'},
    'Lopez 2018':     {'region': 'Spain',      'rel': 2, 'exclude': False,
                       'note': 'RDI/SDI comparison; WP_ET = 0.23 kg/m³ reported'},
    'Egea 2010':      {'region': 'Spain',      'rel': 2, 'exclude': False,
                       'note': 'PRD and RDI; 3 seasons; cv Marta'},
    'Girona 2005':    {'region': 'Spain',      'rel': 2, 'exclude': False,
                       'note': 'RDI kernel-filling; 4-yr; R² low (0.39) — noisy response'},
    'Egea 2012':      {'region': 'Spain',      'rel': 1, 'exclude': True,
                       'note': '⚠ EXCLUDED: young orchard (6-yr cumul ÷6); tree age confounds water effect'},
    'Miras 2023':     {'region': 'Spain (meta)','rel': 0, 'exclude': True,
                       'note': '⚠ EXCLUDED: meta-analysis of ~15 Spanish studies — double-counts Spain'},
}

all_sources = (miras_2023 + egea_2010 + egea_2012 + girona_2005 + lopez_2018 +
               moldero_2021 + goldhamer_2005 + goldhamer_2016 + naor_2017 + sperling_2025)

df = pd.DataFrame(all_sources, columns=COLS)
df['Irrigation_m3_per_dunam'] = pd.to_numeric(df['Irrigation_m3_per_dunam'])
df['Yield_kg_per_dunam']      = pd.to_numeric(df['Yield_kg_per_dunam'])
df['Seasonal_Rainfall_mm']    = pd.to_numeric(df['Seasonal_Rainfall_mm'])
df = df.sort_values('Irrigation_m3_per_dunam').reset_index(drop=True)

# Add metadata columns
df['Region']      = df['Study'].map(lambda s: STUDY_META[s]['region'])
df['Relevance']   = df['Study'].map(lambda s: STUDY_META[s]['rel'])
df['Use_in_fit']  = df['Study'].map(lambda s: not STUDY_META[s]['exclude'])

df_fit = df[df['Use_in_fit']].copy()

print(f"Total observations : {len(df)}")
print(f"Used for fitting   : {len(df_fit)} points from {df_fit['Study'].nunique()} studies")
print(f"Excluded (flagged) : {len(df) - len(df_fit)} points\n")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — PARAMETER EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

print("=" * 62)
print("STEP 1 — Per-study linear fit  (Y = a·x + b)")
print("=" * 62)

study_params = {}
for study in df_fit['Study'].unique():
    grp  = df_fit[df_fit['Study'] == study]
    x, y = grp['Irrigation_m3_per_dunam'].values, grp['Yield_kg_per_dunam'].values
    n    = len(grp)
    if n < 3:
        study_params[study] = dict(slope=np.nan, intercept=np.nan, r2=np.nan, n=n)
        print(f"  {study:<20s}  n={n} — too few points for reliable fit, skipped")
        continue
    a, b   = np.polyfit(x, y, 1)
    y_pred = a * x + b
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - y.mean()) ** 2)
    r2     = 1 - ss_res / ss_tot if ss_tot > 0 else 0
    rel    = STUDY_META[study]['rel']
    study_params[study] = dict(slope=a, intercept=b, r2=r2, n=n, rel=rel,
                               region=STUDY_META[study]['region'])
    print(f"  {study:<20s}  slope={a:+.4f} kg/m³  intercept={b:+.1f}  R²={r2:.2f}  "
          f"n={n}  rel={rel}  [{STUDY_META[study]['region']}]")

# ── STEP 2: Weighted ensemble ─────────────────────────────────────────────────
print("\n" + "=" * 62)
print("STEP 2 — Relevance-weighted ensemble")
print("  weight_i = relevance score (1–5)")
print("  slope_prior  = Σ(w_i · slope_i)  / Σ(w_i)")
print("  slope_SD     = √( Σ(w_i · (slope_i − slope_prior)²) / Σ(w_i) )")
print("=" * 62)

valid = {s: v for s, v in study_params.items() if not np.isnan(v['slope'])}
w     = np.array([v['rel']       for v in valid.values()], dtype=float)
sl    = np.array([v['slope']     for v in valid.values()])
ic    = np.array([v['intercept'] for v in valid.values()])

slope_prior  = np.dot(w, sl) / w.sum()
inter_prior  = np.dot(w, ic) / w.sum()
slope_sd     = np.sqrt(np.dot(w, (sl - slope_prior) ** 2) / w.sum())
inter_sd     = np.sqrt(np.dot(w, (ic - inter_prior) ** 2) / w.sum())

print(f"\n  Slope  prior : {slope_prior:.4f} ± {slope_sd:.4f} kg/m³")
print(f"  Intercept prior: {inter_prior:.1f} ± {inter_sd:.1f} kg/dunam")

print("\n  Contribution of each study to ensemble:")
for study, v in valid.items():
    contribution = (v['rel'] / w.sum()) * 100
    print(f"    {study:<20s}  weight={v['rel']}  share={contribution:.1f}%  "
          f"slope={v['slope']:.4f}")

# ── STEP 3: Mitscherlich saturation fit ───────────────────────────────────────
print("\n" + "=" * 62)
print("STEP 3 — Mitscherlich saturation fit (pooled)")
print("  Y = Y_max · (1 − exp(−k · (x − x0)))")
print("=" * 62)

def mitscherlich(x, Y_max, k, x0):
    return Y_max * (1 - np.exp(-k * (x - x0)))

x_all = df_fit['Irrigation_m3_per_dunam'].values
y_all = df_fit['Yield_kg_per_dunam'].values

try:
    popt, pcov = curve_fit(
        mitscherlich, x_all, y_all,
        p0=[350, 0.003, 50],
        bounds=([100, 0.0001, -200], [600, 0.05, 400]),
        maxfev=10000
    )
    perr   = np.sqrt(np.diag(pcov))
    y_pred = mitscherlich(x_all, *popt)
    rmse   = np.sqrt(np.mean((y_all - y_pred) ** 2))
    r2_m   = 1 - np.sum((y_all - y_pred)**2) / np.sum((y_all - y_all.mean())**2)
    mits_ok = True
    print(f"\n  Y_max = {popt[0]:.1f} ± {perr[0]:.1f} kg/dunam")
    print(f"  k     = {popt[1]:.5f} ± {perr[1]:.5f}  (curvature)")
    print(f"  x0    = {popt[2]:.1f} ± {perr[2]:.1f} m³/dunam  (x-offset)")
    print(f"  RMSE  = {rmse:.1f} kg/dunam   R² = {r2_m:.3f}")
    print(f"\n  Full equation:")
    print(f"  Y = {popt[0]:.1f} · (1 − exp(−{popt[1]:.5f} · (x − {popt[2]:.1f})))")
except Exception as e:
    mits_ok = False
    print(f"  Fit failed: {e}")

# ── STEP 4: Bayesian priors summary ───────────────────────────────────────────
print("\n" + "=" * 62)
print("STEP 4 — Bayesian priors for farm calibration")
print("=" * 62)
print(f"  slope     ~ Normal({slope_prior:.4f},  {slope_sd:.4f}²)  kg/m³")
print(f"  intercept ~ Normal({inter_prior:.1f},  {inter_sd:.1f}²)  kg/dunam")
if mits_ok:
    print(f"  Y_max     ~ Normal({popt[0]:.1f},  {perr[0]:.1f}²)   kg/dunam")
    print(f"  k         ~ HalfNormal({popt[1]:.5f})")
    print(f"  x0        ~ Normal({popt[2]:.1f},  {perr[2]:.1f}²)   m³/dunam")

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

THIN   = Side(style='thin',   color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _style(cell, val=None, bold=False, fill_hex=None, font_color='000000',
           italic=False, align='center', wrap=False, size=10):
    if val is not None:
        cell.value = val
    cell.font = Font(name='Arial', bold=bold, italic=italic, size=size,
                     color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = BORDER
    if fill_hex:
        cell.fill = PatternFill('solid', start_color=fill_hex)

def hdr(cell, val):
    _style(cell, val, bold=True, fill_hex='1F4E79', font_color='FFFFFF', wrap=True)

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════
# Sheet 1 — All_Observations
# ═══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'All_Observations'

ws1.merge_cells('A1:H1')
_style(ws1['A1'],
       'Almond Literature — Irrigation–Yield Data  '
       '(all yields: kg kernel/dunam/year  |  irrigation: m³/dunam)',
       bold=True, size=12, font_color='1F4E79')
ws1.row_dimensions[1].height = 24

ws1.merge_cells('A2:H2')
_style(ws1['A2'],
       'Orange fill = unit correction applied at source  |  '
       'Yellow fill = excluded from parameter fitting  |  '
       'Green fill = Israel studies (highest relevance)',
       italic=True, size=9, font_color='595959')
ws1.row_dimensions[2].height = 16

obs_headers = ['Irrigation_m3_per_dunam', 'Yield_kg_per_dunam',
               'Seasonal_Rainfall_mm', 'Cultivar', 'Study',
               'Region', 'Relevance', 'Notes']
for c, h in enumerate(obs_headers, 1):
    hdr(ws1.cell(row=3, column=c), h)
ws1.row_dimensions[3].height = 36

FILL_CORRECTED = 'FCE4D6'   # orange — unit correction applied
FILL_EXCLUDED  = 'FFF2CC'   # yellow — excluded from fit
FILL_ISRAEL    = 'E2EFDA'   # green  — high relevance
FILL_NORMAL    = None

corrected_studies = {'Naor 2017', 'Goldhamer 2005', 'Egea 2012'}

for r_idx, row in df.iterrows():
    excel_row = r_idx + 4
    study     = row['Study']
    excluded  = STUDY_META[study]['exclude']
    corrected = study in corrected_studies
    israeli   = STUDY_META[study]['region'] == 'Israel'

    fill = (FILL_EXCLUDED  if excluded  else
            FILL_CORRECTED if corrected else
            FILL_ISRAEL    if israeli   else
            FILL_NORMAL)

    values = [row['Irrigation_m3_per_dunam'], row['Yield_kg_per_dunam'],
              row['Seasonal_Rainfall_mm'],    row['Cultivar'],
              study, row['Region'],           row['Relevance'],
              row['Notes'] if pd.notna(row['Notes']) else '']
    for c, val in enumerate(values, 1):
        _style(ws1.cell(excel_row, c), val, fill_hex=fill)

obs_widths = [28, 24, 24, 14, 18, 16, 12, 52]
for i, w in enumerate(obs_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════
# Sheet 2 — Study_Summaries
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Study_Summaries')

ws2.merge_cells('A1:I1')
_style(ws2['A1'], 'Almond — Study Summaries', bold=True, size=12, font_color='1F4E79')
ws2.row_dimensions[1].height = 24

sum_headers = ['Study', 'Region', 'Relevance\n(0–5)', 'N points',
               'Irrigation range\n(m³/dunam)', 'Yield range\n(kg/dunam)',
               'Max yield\n(kg/dunam)', 'Used in fit', 'Notes']
for c, h in enumerate(sum_headers, 1):
    hdr(ws2.cell(row=2, column=c), h)
ws2.row_dimensions[2].height = 38

for r_idx, study in enumerate(sorted(STUDY_META.keys(),
                               key=lambda s: -STUDY_META[s]['rel']), 3):
    m    = STUDY_META[study]
    grp  = df[df['Study'] == study]
    fill = (FILL_EXCLUDED if m['exclude'] else
            FILL_ISRAEL   if m['region'] == 'Israel' else FILL_NORMAL)

    row_vals = [
        study,
        m['region'],
        m['rel'],
        len(grp),
        f"{grp['Irrigation_m3_per_dunam'].min():.0f}–{grp['Irrigation_m3_per_dunam'].max():.0f}",
        f"{grp['Yield_kg_per_dunam'].min():.1f}–{grp['Yield_kg_per_dunam'].max():.1f}",
        round(grp['Yield_kg_per_dunam'].max(), 1),
        'No ⚠' if m['exclude'] else 'Yes',
        m['note'],
    ]
    for c, val in enumerate(row_vals, 1):
        _style(ws2.cell(r_idx, c), val, fill_hex=fill,
               align='left' if c == 9 else 'center', wrap=(c == 9))
    ws2.row_dimensions[r_idx].height = 36

sum_widths = [18, 16, 14, 10, 24, 22, 18, 12, 62]
for i, w in enumerate(sum_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════════════
# Sheet 3 — Parameter_Extraction
# ═══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Parameter_Extraction')

ws3.merge_cells('A1:G1')
_style(ws3['A1'], 'Almond — Production Function Parameter Extraction',
       bold=True, size=12, font_color='1F4E79')
ws3.row_dimensions[1].height = 24

# ── 3a: Per-study fits ────────────────────────────────────────────────
ws3.merge_cells('A2:G2')
_style(ws3['A2'], 'STEP 1 — Per-study linear fits:  Y = a·x + b  '
       '(fitting studies only; n ≥ 3 required)',
       bold=True, fill_hex='DEEAF1', font_color='1F4E79', align='left')

fit_headers = ['Study', 'Region', 'Relevance', 'N points',
               'Slope a (kg/m³)', 'Intercept b (kg/dunam)', 'R²']
for c, h in enumerate(fit_headers, 1):
    hdr(ws3.cell(row=3, column=c), h)

r = 4
for study in sorted(study_params.keys(), key=lambda s: -STUDY_META[s]['rel']):
    v    = study_params[study]
    fill = FILL_ISRAEL if STUDY_META[study]['region'] == 'Israel' else FILL_NORMAL
    row_vals = [study,
                STUDY_META[study]['region'],
                STUDY_META[study]['rel'],
                v['n'],
                round(v['slope'], 4)     if not np.isnan(v['slope']) else 'n/a',
                round(v['intercept'], 1) if not np.isnan(v['slope']) else 'n/a',
                round(v['r2'], 3)        if not np.isnan(v['slope']) else 'n/a']
    for c, val in enumerate(row_vals, 1):
        _style(ws3.cell(r, c), val, fill_hex=fill)
    r += 1

# ── 3b: Weighted ensemble ─────────────────────────────────────────────
r += 1
ws3.merge_cells(f'A{r}:G{r}')
_style(ws3.cell(r, 1),
       'STEP 2 — Relevance-weighted ensemble   '
       'weight_i = relevance score   '
       'slope_prior = Σ(w·slope)/Σ(w)   '
       'slope_SD = √(Σ(w·(slope−prior)²)/Σ(w))',
       bold=True, fill_hex='DEEAF1', font_color='1F4E79', align='left', wrap=True)
ws3.row_dimensions[r].height = 32
r += 1

ens_headers = ['Parameter', 'Prior mean', 'Prior SD', 'Distribution', '',  '', 'Interpretation']
for c, h in enumerate(ens_headers, 1):
    hdr(ws3.cell(r, c), h)
r += 1
ens_rows = [
    ('Slope a  (kg/m³)',       round(slope_prior, 4), round(slope_sd, 4),
     'Normal(μ, σ²)', '', '',
     'dY/dIrrigation — marginal yield per m³'),
    ('Intercept b  (kg/dunam)', round(inter_prior, 1), round(inter_sd, 1),
     'Normal(μ, σ²)', '', '',
     'Yield at zero irrigation (rainfall-supported baseline)'),
]
for vals in ens_rows:
    for c, val in enumerate(vals, 1):
        _style(ws3.cell(r, c), val, fill_hex='F2F7FB')
    r += 1

# ── 3c: Mitscherlich fit ──────────────────────────────────────────────
r += 1
ws3.merge_cells(f'A{r}:G{r}')
_style(ws3.cell(r, 1),
       'STEP 3 — Mitscherlich saturation fit (pooled):  '
       'Y = Y_max · (1 − exp(−k · (x − x0)))',
       bold=True, fill_hex='DEEAF1', font_color='1F4E79', align='left')
r += 1

mits_headers = ['Parameter', 'Estimate', 'Std Error', 'Distribution', '', '', 'Interpretation']
for c, h in enumerate(mits_headers, 1):
    hdr(ws3.cell(r, c), h)
r += 1
if mits_ok:
    mits_rows = [
        ('Y_max  (kg/dunam)', round(popt[0], 1), round(perr[0], 1),
         'Normal(μ, σ²)', '', '', 'Asymptotic maximum yield'),
        ('k  (curvature)',    round(popt[1], 5), round(perr[1], 5),
         'HalfNormal', '',  '',  'Rate of diminishing returns'),
        ('x0  (m³/dunam)',    round(popt[2], 1), round(perr[2], 1),
         'Normal(μ, σ²)', '', '', 'Irrigation offset (x-axis shift)'),
        ('RMSE  (kg/dunam)',  round(rmse, 1),    '—', '—', '', '',
         f'R² = {r2_m:.3f}'),
    ]
    for vals in mits_rows:
        for c, val in enumerate(vals, 1):
            _style(ws3.cell(r, c), val, fill_hex='F2F7FB')
        r += 1

    r += 1
    ws3.merge_cells(f'A{r}:G{r}')
    _style(ws3.cell(r, 1),
           f'Full equation:  Y = {popt[0]:.1f} × (1 − exp(−{popt[1]:.5f} × (x − {popt[2]:.1f})))',
           bold=True, fill_hex='E2EFDA', font_color='375623', align='left')
    r += 1

# ── 3d: Study weight table ────────────────────────────────────────────
r += 1
ws3.merge_cells(f'A{r}:G{r}')
_style(ws3.cell(r, 1),
       'STEP 2 detail — Per-study contribution to weighted ensemble',
       bold=True, fill_hex='DEEAF1', font_color='1F4E79', align='left')
r += 1

wt_headers = ['Study', 'Relevance weight', 'Weight share (%)',
              'Slope (kg/m³)', 'Deviation from prior', '', 'Weighted contribution']
for c, h in enumerate(wt_headers, 1):
    hdr(ws3.cell(r, c), h)
r += 1

w_np = np.array([v["rel"] for v in valid.values()], dtype=float)
sl_np = np.array([v["slope"] for v in valid.values()])
w_total = w_np.sum()
for study, wi, sli in zip(valid.keys(), w_np, sl_np):
    share = wi / w_total * 100
    dev   = sli - slope_prior
    contrib = wi * sli / w_total
    fill = FILL_ISRAEL if STUDY_META[study]['region'] == 'Israel' else FILL_NORMAL
    for c, val in enumerate([study, wi, round(share,1), round(sli,4),
                              round(dev,4), '', round(contrib,4)], 1):
        _style(ws3.cell(r, c), val, fill_hex=fill)
    r += 1

# ── Final prior summary row ────────────────────────────────────────────
r += 1
ws3.merge_cells(f'A{r}:G{r}')
_style(ws3.cell(r, 1),
       f'FINAL PRIORS:   '
       f'slope ~ N({slope_prior:.4f}, {slope_sd:.4f}²)  kg/m³   |   '
       f'intercept ~ N({inter_prior:.1f}, {inter_sd:.1f}²)  kg/dunam'
       + (f'   |   Y_max ~ N({popt[0]:.1f}, {perr[0]:.1f}²)  kg/dunam' if mits_ok else ''),
       bold=True, fill_hex='1F4E79', font_color='FFFFFF', align='left', wrap=True)
ws3.row_dimensions[r].height = 28

par_widths = [22, 18, 18, 14, 20, 18, 42]
for i, w_col in enumerate(par_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w_col

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
out_path = 'almond_literature_calibration_data.xlsx'
wb.save(out_path)
print(f"\n✅  Saved → {out_path}")
print(f"    Sheets: All_Observations ({len(df)} rows) | "
      f"Study_Summaries ({len(STUDY_META)} studies) | Parameter_Extraction")